#!/usr/bin/env python3
"""Convert the manually maintained Middle East strike Excel workbook to JSON.

Expected workbook sheet: Események
The script intentionally supports the current Hungarian column names, so the
existing workbook can be used without restructuring it.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import sys
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

EXPECTED_HEADERS = {
    "Dátum",
    "Támadó fél",
    "Célország",
    "Helyszín / célpont",
    "Szélesség",
    "Hosszúság",
    "Esemény leírása",
    "Támadás típusa",
    "Koordináta / adat megjegyzés",
    "Bizonyosság",
    "Forrás URL",
}

ATTACKER_MAP = {
    "usa": "USA",
    "egyesült államok": "USA",
    "united states": "USA",
    "irán": "IRAN",
    "iran": "IRAN",
}

CONFIDENCE_MAP = {
    "magas": "HIGH",
    "közepes": "MEDIUM",
    "alacsony": "LOW",
    "high": "HIGH",
    "medium": "MEDIUM",
    "low": "LOW",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y.%m.%d.", "%d.%m.%Y", "%d.%m.%Y."):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"Nem értelmezhető dátum: {value!r}")


def normalize_attacker(value: Any) -> str:
    text = clean_text(value)
    return ATTACKER_MAP.get(text.casefold(), text.upper())


def normalize_confidence(value: Any) -> str:
    text = clean_text(value)
    return CONFIDENCE_MAP.get(text.casefold(), text.upper() or "UNKNOWN")


def normalize_coordinate(value: Any, field_name: str) -> float:
    if value is None or value == "":
        raise ValueError(f"Hiányzó koordináta: {field_name}")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Hibás koordináta ({field_name}): {value!r}") from exc
    if not math.isfinite(number):
        raise ValueError(f"Nem véges koordináta ({field_name}): {value!r}")
    if field_name == "latitude" and not -90 <= number <= 90:
        raise ValueError(f"Szélesség tartományon kívül: {number}")
    if field_name == "longitude" and not -180 <= number <= 180:
        raise ValueError(f"Hosszúság tartományon kívül: {number}")
    return round(number, 6)


def make_event_id(event_date: str, attacker: str, country: str, location: str, row_number: int) -> str:
    raw = f"{event_date}|{attacker}|{country}|{location}|{row_number}"
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:8].upper()
    attacker_code = "US" if attacker == "USA" else "IR" if attacker == "IRAN" else "OT"
    return f"ME-{event_date.replace('-', '')}-{attacker_code}-{digest}"


def read_events(input_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    workbook = load_workbook(input_path, data_only=True, read_only=True)
    if "Események" not in workbook.sheetnames:
        raise RuntimeError("Az Excel nem tartalmaz 'Események' nevű munkalapot.")

    sheet = workbook["Események"]
    headers = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    missing = sorted(EXPECTED_HEADERS - set(headers))
    if missing:
        raise RuntimeError("Hiányzó kötelező oszlopok: " + ", ".join(missing))

    column = {name: index for index, name in enumerate(headers)}
    events: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in row):
            continue
        try:
            event_date = normalize_date(row[column["Dátum"]])
            attacker = normalize_attacker(row[column["Támadó fél"]])
            country = clean_text(row[column["Célország"]])
            location = clean_text(row[column["Helyszín / célpont"]])
            latitude = normalize_coordinate(row[column["Szélesség"]], "latitude")
            longitude = normalize_coordinate(row[column["Hosszúság"]], "longitude")
            description = clean_text(row[column["Esemény leírása"]])
            strike_type = clean_text(row[column["Támadás típusa"]])
            coordinate_note = clean_text(row[column["Koordináta / adat megjegyzés"]])
            confidence_hu = clean_text(row[column["Bizonyosság"]])
            source_url = clean_text(row[column["Forrás URL"]])

            required_text = {
                "Támadó fél": attacker,
                "Célország": country,
                "Helyszín / célpont": location,
                "Esemény leírása": description,
                "Támadás típusa": strike_type,
                "Forrás URL": source_url,
            }
            empty_fields = [name for name, value in required_text.items() if not value]
            if empty_fields:
                raise ValueError("Hiányzó kötelező mező(k): " + ", ".join(empty_fields))

            if attacker not in {"USA", "IRAN"}:
                warnings.append({"row": row_number, "warning": f"Ismeretlen támadó fél: {attacker}"})
            if source_url and not source_url.startswith(("http://", "https://")):
                warnings.append({"row": row_number, "warning": "A forrás URL nem http/https címmel kezdődik."})

            event = {
                "event_id": make_event_id(event_date, attacker, country, location, row_number),
                "date": event_date,
                "attacker": attacker,
                "target_country": country,
                "target_location": location,
                "latitude": latitude,
                "longitude": longitude,
                "description": description,
                "strike_type": strike_type,
                "coordinate_note": coordinate_note,
                "confidence": normalize_confidence(confidence_hu),
                "confidence_label_hu": confidence_hu,
                "source_url": source_url,
                "source_row": row_number,
            }
            events.append(event)
        except Exception as exc:
            warnings.append({"row": row_number, "error": str(exc)})

    events.sort(key=lambda item: (item["date"], item["attacker"], item["event_id"]))
    return events, warnings


def build_summary(events: list[dict[str, Any]]) -> dict[str, Any]:
    by_attacker = Counter(event["attacker"] for event in events)
    by_country = Counter(event["target_country"] for event in events)
    by_date = Counter(event["date"] for event in events)
    dates = [event["date"] for event in events]
    return {
        "event_count": len(events),
        "date_start": min(dates) if dates else None,
        "date_end": max(dates) if dates else None,
        "attackers": dict(sorted(by_attacker.items())),
        "target_countries": dict(sorted(by_country.items(), key=lambda x: (-x[1], x[0]))),
        "daily_counts": dict(sorted(by_date.items())),
    }


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="USA–Irán csapásadatok Excel → JSON feldolgozó")
    parser.add_argument("--input", default="data/manual/usa_iran_tamadasok.xlsx")
    parser.add_argument("--output", default="data/strike_history.json")
    parser.add_argument("--summary", default="data/strike_history_summary.json")
    parser.add_argument("--validation", default="data/strike_history_validation.json")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"HIBA: Nem található a bemeneti Excel: {input_path}", file=sys.stderr)
        return 1

    events, warnings = read_events(input_path)
    errors = [item for item in warnings if "error" in item]
    generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

    dataset = {
        "generated_at": generated_at,
        "dataset": "middle_east_strike_history",
        "source_file": input_path.name,
        "summary": build_summary(events),
        "events": events,
    }
    write_json(Path(args.output), dataset)
    write_json(Path(args.summary), {"generated_at": generated_at, **build_summary(events)})
    write_json(Path(args.validation), {
        "generated_at": generated_at,
        "source_file": input_path.name,
        "valid_event_count": len(events),
        "error_count": len(errors),
        "warning_count": len(warnings) - len(errors),
        "items": warnings,
    })

    print(f"Feldolgozott események: {len(events)}")
    print(f"Hibák: {len(errors)} | Figyelmeztetések: {len(warnings) - len(errors)}")
    if errors:
        print("HIBA: Egy vagy több Excel-sor nem volt feldolgozható.", file=sys.stderr)
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
